"""Export simulation results to CSV, JSON, SARIF, and Excel formats."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

from infrasim.simulator.engine import SimulationReport

logger = logging.getLogger(__name__)


def _report_rows(report: SimulationReport) -> list[dict]:
    """Flatten a SimulationReport into a list of row dicts for CSV export."""
    rows: list[dict] = []
    for result in report.results:
        base = {
            "scenario_id": result.scenario.id,
            "scenario_name": result.scenario.name,
            "scenario_description": result.scenario.description,
            "risk_score": round(result.risk_score, 4),
            "is_critical": result.is_critical,
            "is_warning": result.is_warning,
            "cascade_trigger": result.cascade.trigger,
            "cascade_severity": round(result.cascade.severity, 4),
            "affected_components": len(result.cascade.effects),
        }

        if result.cascade.effects:
            for effect in result.cascade.effects:
                row = {
                    **base,
                    "component_id": effect.component_id,
                    "component_name": effect.component_name,
                    "health": effect.health.value,
                    "reason": effect.reason,
                    "estimated_time_seconds": effect.estimated_time_seconds,
                }
                rows.append(row)
        else:
            # Scenario with no cascade effects still gets a row
            rows.append({
                **base,
                "component_id": "",
                "component_name": "",
                "health": "",
                "reason": "",
                "estimated_time_seconds": 0,
            })

    return rows


def export_csv(report: SimulationReport, path: Path) -> Path:
    """Export simulation results as a CSV file.

    Args:
        report: The simulation report to export.
        path: Destination file path.

    Returns:
        The resolved Path of the written file.
    """
    path = Path(path)
    rows = _report_rows(report)

    if not rows:
        # Write an empty CSV with just headers
        fieldnames = [
            "scenario_id", "scenario_name", "scenario_description",
            "risk_score", "is_critical", "is_warning",
            "cascade_trigger", "cascade_severity", "affected_components",
            "component_id", "component_name", "health", "reason",
            "estimated_time_seconds",
        ]
    else:
        fieldnames = list(rows[0].keys())

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return path.resolve()


def _report_to_export_dict(report: SimulationReport) -> dict:
    """Build a JSON-serialisable dict from a SimulationReport."""

    def _effect_dict(e):
        return {
            "component_id": e.component_id,
            "component_name": e.component_name,
            "health": e.health.value,
            "reason": e.reason,
            "estimated_time_seconds": e.estimated_time_seconds,
            "metrics_impact": e.metrics_impact,
        }

    def _result_dict(r):
        return {
            "scenario_id": r.scenario.id,
            "scenario_name": r.scenario.name,
            "scenario_description": r.scenario.description,
            "risk_score": round(r.risk_score, 4),
            "is_critical": r.is_critical,
            "is_warning": r.is_warning,
            "cascade": {
                "trigger": r.cascade.trigger,
                "severity": round(r.cascade.severity, 4),
                "effects": [_effect_dict(e) for e in r.cascade.effects],
            },
        }

    return {
        "resilience_score": round(report.resilience_score, 2),
        "total_scenarios": len(report.results),
        "critical_count": len(report.critical_findings),
        "warning_count": len(report.warnings),
        "passed_count": len(report.passed),
        "results": [_result_dict(r) for r in report.results],
    }


def export_json(report: SimulationReport, path: Path) -> Path:
    """Export simulation results as a formatted JSON file.

    Args:
        report: The simulation report to export.
        path: Destination file path.

    Returns:
        The resolved Path of the written file.
    """
    path = Path(path)
    data = _report_to_export_dict(report)

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)

    return path.resolve()


# ---------------------------------------------------------------------------
# SARIF export (GitHub / GitLab security tab)
# ---------------------------------------------------------------------------

def export_sarif(report: SimulationReport) -> str:
    """Export simulation results in SARIF 2.1.0 format.

    SARIF (Static Analysis Results Interchange Format) is understood by
    GitHub Advanced Security, GitLab SAST, and other CI/CD platforms.

    Args:
        report: The simulation report to export.

    Returns:
        A JSON string in SARIF 2.1.0 format.
    """
    rules: list[dict] = []
    results_list: list[dict] = []
    seen_rule_ids: set[str] = set()

    for sr in report.results:
        # Only include critical and warning findings
        if not sr.is_critical and not sr.is_warning:
            continue

        rule_id = sr.scenario.id
        if rule_id not in seen_rule_ids:
            seen_rule_ids.add(rule_id)
            level = "error" if sr.is_critical else "warning"
            rules.append({
                "id": rule_id,
                "name": sr.scenario.name,
                "shortDescription": {
                    "text": sr.scenario.name,
                },
                "fullDescription": {
                    "text": sr.scenario.description or sr.scenario.name,
                },
                "defaultConfiguration": {
                    "level": level,
                },
            })

        sarif_level = "error" if sr.is_critical else "warning"

        # Build message from cascade effects
        effect_messages = []
        for effect in sr.cascade.effects:
            effect_messages.append(
                f"{effect.component_name}: {effect.health.value} - {effect.reason}"
            )
        message_text = (
            f"{sr.scenario.description or sr.scenario.name} "
            f"(risk_score={sr.risk_score:.1f})"
        )
        if effect_messages:
            message_text += "\nAffected: " + "; ".join(effect_messages[:5])

        result_entry: dict = {
            "ruleId": rule_id,
            "level": sarif_level,
            "message": {
                "text": message_text,
            },
            "properties": {
                "risk_score": round(sr.risk_score, 4),
                "cascade_severity": round(sr.cascade.severity, 4),
                "affected_components": len(sr.cascade.effects),
            },
        }
        results_list.append(result_entry)

    sarif = {
        "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/main/sarif-2.1/schema/sarif-schema-2.1.0.json",
        "version": "2.1.0",
        "runs": [
            {
                "tool": {
                    "driver": {
                        "name": "FaultRay",
                        "informationUri": "https://github.com/mattyopon/infrasim",
                        "version": "2.1.0",
                        "rules": rules,
                    }
                },
                "results": results_list,
            }
        ],
    }

    return json.dumps(sarif, indent=2, ensure_ascii=False)


def export_sarif_file(report: SimulationReport, path: Path) -> Path:
    """Export SARIF to a file.

    Args:
        report: The simulation report to export.
        path: Destination file path.

    Returns:
        The resolved Path of the written file.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    sarif_str = export_sarif(report)
    path.write_text(sarif_str, encoding="utf-8")
    return path.resolve()


# ---------------------------------------------------------------------------
# Excel export (optional dependency: openpyxl)
# ---------------------------------------------------------------------------

def export_excel(report: SimulationReport, path: Path) -> Path:
    """Export simulation results as an Excel (.xlsx) file.

    Requires the ``openpyxl`` package.  If it is not installed, an
    ``ImportError`` is raised with a helpful message.

    Args:
        report: The simulation report to export.
        path: Destination file path.

    Returns:
        The resolved Path of the written file.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, size=14)
    ws_summary["A1"] = "FaultRay Simulation Report"
    ws_summary["A1"].font = header_font

    ws_summary["A3"] = "Resilience Score"
    ws_summary["B3"] = round(report.resilience_score, 1)
    ws_summary["A4"] = "Total Scenarios"
    ws_summary["B4"] = len(report.results)
    ws_summary["A5"] = "Critical"
    ws_summary["B5"] = len(report.critical_findings)
    ws_summary["A6"] = "Warning"
    ws_summary["B6"] = len(report.warnings)
    ws_summary["A7"] = "Passed"
    ws_summary["B7"] = len(report.passed)

    for row in range(3, 8):
        ws_summary.cell(row=row, column=1).font = Font(bold=True)

    # --- Results sheet ---
    ws_results = wb.create_sheet("Results")
    rows = _report_rows(report)

    if rows:
        headers = list(rows[0].keys())
    else:
        headers = [
            "scenario_id", "scenario_name", "scenario_description",
            "risk_score", "is_critical", "is_warning",
            "cascade_trigger", "cascade_severity", "affected_components",
            "component_id", "component_name", "health", "reason",
            "estimated_time_seconds",
        ]

    # Write headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=header)
        cell.font = header_text_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_results.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Conditional formatting by severity
        is_critical = row_data.get("is_critical", False)
        is_warning = row_data.get("is_warning", False)
        if is_critical:
            for col_idx in range(1, len(headers) + 1):
                ws_results.cell(row=row_idx, column=col_idx).fill = critical_fill
        elif is_warning:
            for col_idx in range(1, len(headers) + 1):
                ws_results.cell(row=row_idx, column=col_idx).fill = warning_fill

    # Auto-size columns (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, min(len(rows) + 2, 50)):
            val = ws_results.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws_results.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max_len + 2

    wb.save(str(path))
    return path.resolve()
