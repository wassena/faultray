"""Tests for SARIF and Excel export formats."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from infrasim.model.components import Component, ComponentType, HealthStatus
from infrasim.model.graph import InfraGraph
from infrasim.reporter.export import (
    _report_rows,
    _report_to_export_dict,
    export_csv,
    export_excel,
    export_json,
    export_sarif,
    export_sarif_file,
)
from infrasim.simulator.cascade import CascadeChain, CascadeEffect
from infrasim.simulator.engine import ScenarioResult, SimulationEngine, SimulationReport
from infrasim.simulator.scenarios import Scenario


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_graph() -> InfraGraph:
    """Build a minimal graph for testing."""
    graph = InfraGraph()
    graph.add_component(
        Component(id="web", name="Web Server", type=ComponentType.WEB_SERVER)
    )
    graph.add_component(
        Component(id="db", name="Database", type=ComponentType.DATABASE)
    )
    return graph


@pytest.fixture
def sample_report(sample_graph: InfraGraph) -> SimulationReport:
    """Build a SimulationReport with mixed severity scenarios."""
    # Critical scenario
    critical_scenario = Scenario(
        id="s-critical-1",
        name="DB Total Failure",
        description="Primary database goes offline",
        faults=[],
    )
    critical_cascade = CascadeChain(trigger="DB Total Failure", total_components=2)
    critical_cascade.effects.append(
        CascadeEffect(
            component_id="db",
            component_name="Database",
            health=HealthStatus.DOWN,
            reason="Primary node unreachable",
            estimated_time_seconds=30,
        )
    )
    critical_cascade.likelihood = 0.8

    # Warning scenario
    warning_scenario = Scenario(
        id="s-warning-1",
        name="Cache Degradation",
        description="Cache hit rate drops significantly",
        faults=[],
    )
    warning_cascade = CascadeChain(trigger="Cache Degradation", total_components=2)
    warning_cascade.effects.append(
        CascadeEffect(
            component_id="web",
            component_name="Web Server",
            health=HealthStatus.DEGRADED,
            reason="Increased latency due to cache miss",
            estimated_time_seconds=10,
        )
    )
    warning_cascade.likelihood = 0.3

    # Passed scenario
    passed_scenario = Scenario(
        id="s-passed-1",
        name="Minor Network Jitter",
        description="Slight increase in network latency",
        faults=[],
    )
    passed_cascade = CascadeChain(trigger="Minor Network Jitter", total_components=2)

    results = [
        ScenarioResult(scenario=critical_scenario, cascade=critical_cascade, risk_score=8.5),
        ScenarioResult(scenario=warning_scenario, cascade=warning_cascade, risk_score=5.0),
        ScenarioResult(scenario=passed_scenario, cascade=passed_cascade, risk_score=1.0),
    ]

    return SimulationReport(
        results=results,
        resilience_score=65.0,
        total_generated=3,
    )


# ---------------------------------------------------------------------------
# SARIF export tests
# ---------------------------------------------------------------------------


class TestSarifExport:
    def test_sarif_valid_json(self, sample_report: SimulationReport):
        """export_sarif() should return valid JSON."""
        sarif_str = export_sarif(sample_report)
        data = json.loads(sarif_str)
        assert data["version"] == "2.1.0"

    def test_sarif_schema(self, sample_report: SimulationReport):
        """SARIF output should have the correct schema structure."""
        data = json.loads(export_sarif(sample_report))
        assert "$schema" in data
        assert "runs" in data
        assert len(data["runs"]) == 1

        run = data["runs"][0]
        assert "tool" in run
        assert run["tool"]["driver"]["name"] == "FaultRay"
        assert "rules" in run["tool"]["driver"]
        assert "results" in run

    def test_sarif_includes_critical_and_warning(self, sample_report: SimulationReport):
        """SARIF results should include critical and warning findings only."""
        data = json.loads(export_sarif(sample_report))
        results = data["runs"][0]["results"]

        # Should have 2 results (critical + warning), not the passed one
        assert len(results) == 2
        levels = {r["level"] for r in results}
        assert "error" in levels
        assert "warning" in levels

    def test_sarif_rules_match_results(self, sample_report: SimulationReport):
        """Each SARIF result should reference a defined rule."""
        data = json.loads(export_sarif(sample_report))
        rules = data["runs"][0]["tool"]["driver"]["rules"]
        results = data["runs"][0]["results"]

        rule_ids = {r["id"] for r in rules}
        for result in results:
            assert result["ruleId"] in rule_ids

    def test_sarif_file_export(self, sample_report: SimulationReport, tmp_path: Path):
        """export_sarif_file() should write SARIF to a file."""
        out = tmp_path / "results.sarif"
        path = export_sarif_file(sample_report, out)
        assert path.exists()

        data = json.loads(path.read_text())
        assert data["version"] == "2.1.0"

    def test_sarif_empty_report(self):
        """export_sarif() with no findings should produce valid but empty results."""
        report = SimulationReport(results=[], resilience_score=100.0)
        data = json.loads(export_sarif(report))
        assert data["runs"][0]["results"] == []
        assert data["runs"][0]["tool"]["driver"]["rules"] == []


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------


class TestExcelExport:
    def _has_openpyxl(self) -> bool:
        try:
            import openpyxl
            return True
        except ImportError:
            return False

    def test_excel_export_creates_file(self, sample_report: SimulationReport, tmp_path: Path):
        """export_excel() should create an .xlsx file."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        out = tmp_path / "results.xlsx"
        path = export_excel(sample_report, out)
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_excel_has_two_sheets(self, sample_report: SimulationReport, tmp_path: Path):
        """Excel file should have Summary and Results sheets."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        import openpyxl

        out = tmp_path / "results.xlsx"
        export_excel(sample_report, out)

        wb = openpyxl.load_workbook(str(out))
        assert "Summary" in wb.sheetnames
        assert "Results" in wb.sheetnames

    def test_excel_summary_values(self, sample_report: SimulationReport, tmp_path: Path):
        """Summary sheet should contain correct values."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        import openpyxl

        out = tmp_path / "results.xlsx"
        export_excel(sample_report, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Summary"]
        assert ws["B3"].value == 65.0  # resilience score
        assert ws["B4"].value == 3     # total scenarios
        assert ws["B5"].value == 1     # critical
        assert ws["B6"].value == 1     # warning
        assert ws["B7"].value == 1     # passed

    def test_excel_results_data(self, sample_report: SimulationReport, tmp_path: Path):
        """Results sheet should contain scenario data rows."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        import openpyxl

        out = tmp_path / "results.xlsx"
        export_excel(sample_report, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Results"]
        # Row 1 is header, data starts at row 2
        assert ws.max_row >= 2  # At least header + 1 data row

    def test_excel_empty_report(self, tmp_path: Path):
        """export_excel() should work with an empty report."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        report = SimulationReport(results=[], resilience_score=100.0)
        out = tmp_path / "empty.xlsx"
        path = export_excel(report, out)
        assert path.exists()

    def test_excel_missing_openpyxl(self, sample_report: SimulationReport, tmp_path: Path):
        """export_excel() should raise ImportError with helpful message when openpyxl missing."""
        import importlib
        import sys

        # Temporarily hide openpyxl
        openpyxl_mod = sys.modules.get("openpyxl")
        sys.modules["openpyxl"] = None  # type: ignore

        try:
            with pytest.raises(ImportError, match="openpyxl"):
                export_excel(sample_report, tmp_path / "test.xlsx")
        finally:
            if openpyxl_mod is not None:
                sys.modules["openpyxl"] = openpyxl_mod
            else:
                sys.modules.pop("openpyxl", None)

    def test_excel_creates_parent_dirs(self, sample_report: SimulationReport, tmp_path: Path):
        """export_excel() should create parent directories."""
        if not self._has_openpyxl():
            pytest.skip("openpyxl not installed")

        out = tmp_path / "nested" / "dir" / "results.xlsx"
        path = export_excel(sample_report, out)
        assert path.exists()


# ---------------------------------------------------------------------------
# CSV export tests
# ---------------------------------------------------------------------------


class TestCsvExport:
    def test_csv_export_creates_file(self, sample_report: SimulationReport, tmp_path: Path):
        """export_csv() should create a .csv file."""
        out = tmp_path / "results.csv"
        path = export_csv(sample_report, out)
        assert path.exists()
        assert path.suffix == ".csv"

    def test_csv_has_header_and_data(self, sample_report: SimulationReport, tmp_path: Path):
        """CSV file should have a header row and data rows."""
        import csv as csv_mod
        out = tmp_path / "results.csv"
        export_csv(sample_report, out)

        with open(out, newline="", encoding="utf-8") as fh:
            reader = csv_mod.DictReader(fh)
            rows = list(reader)

        assert len(rows) >= 1
        assert "scenario_id" in rows[0]
        assert "risk_score" in rows[0]
        assert "component_id" in rows[0]

    def test_csv_empty_report(self, tmp_path: Path):
        """export_csv() should work with an empty report."""
        report = SimulationReport(results=[], resilience_score=100.0)
        out = tmp_path / "empty.csv"
        path = export_csv(report, out)
        assert path.exists()

        # Should still have a header row
        with open(out, encoding="utf-8") as fh:
            content = fh.read()
        assert "scenario_id" in content

    def test_csv_creates_parent_dirs(self, sample_report: SimulationReport, tmp_path: Path):
        """export_csv() should create parent directories."""
        out = tmp_path / "nested" / "dir" / "results.csv"
        path = export_csv(sample_report, out)
        assert path.exists()

    def test_csv_scenario_without_effects(self, tmp_path: Path):
        """Scenario with no cascade effects should still produce a row."""
        scenario = Scenario(
            id="s-no-effect", name="No Effect", description="No effects", faults=[],
        )
        cascade = CascadeChain(trigger="No Effect", total_components=1)
        result = ScenarioResult(scenario=scenario, cascade=cascade, risk_score=0.5)
        report = SimulationReport(results=[result], resilience_score=95.0)

        out = tmp_path / "results.csv"
        export_csv(report, out)

        import csv as csv_mod
        with open(out, newline="", encoding="utf-8") as fh:
            reader = csv_mod.DictReader(fh)
            rows = list(reader)

        assert len(rows) == 1
        assert rows[0]["component_id"] == ""


# ---------------------------------------------------------------------------
# JSON export tests
# ---------------------------------------------------------------------------


class TestJsonExport:
    def test_json_export_creates_file(self, sample_report: SimulationReport, tmp_path: Path):
        """export_json() should create a .json file."""
        out = tmp_path / "results.json"
        path = export_json(sample_report, out)
        assert path.exists()

    def test_json_has_expected_structure(self, sample_report: SimulationReport, tmp_path: Path):
        """JSON file should have the expected structure."""
        out = tmp_path / "results.json"
        export_json(sample_report, out)

        data = json.loads(out.read_text())
        assert "resilience_score" in data
        assert data["resilience_score"] == 65.0
        assert data["total_scenarios"] == 3
        assert data["critical_count"] == 1
        assert data["warning_count"] == 1
        assert data["passed_count"] == 1
        assert len(data["results"]) == 3

    def test_json_result_structure(self, sample_report: SimulationReport, tmp_path: Path):
        """Each result in JSON should have scenario and cascade data."""
        out = tmp_path / "results.json"
        export_json(sample_report, out)

        data = json.loads(out.read_text())
        result = data["results"][0]
        assert "scenario_id" in result
        assert "scenario_name" in result
        assert "risk_score" in result
        assert "cascade" in result
        assert "trigger" in result["cascade"]
        assert "effects" in result["cascade"]

    def test_json_empty_report(self, tmp_path: Path):
        """export_json() should work with empty report."""
        report = SimulationReport(results=[], resilience_score=100.0)
        out = tmp_path / "empty.json"
        path = export_json(report, out)
        assert path.exists()

        data = json.loads(out.read_text())
        assert data["total_scenarios"] == 0
        assert data["results"] == []

    def test_json_creates_parent_dirs(self, sample_report: SimulationReport, tmp_path: Path):
        """export_json() should create parent directories."""
        out = tmp_path / "nested" / "dir" / "results.json"
        path = export_json(sample_report, out)
        assert path.exists()


# ---------------------------------------------------------------------------
# Internal helper tests
# ---------------------------------------------------------------------------


class TestReportRows:
    def test_report_rows_with_effects(self, sample_report: SimulationReport):
        """_report_rows should expand cascade effects into rows."""
        rows = _report_rows(sample_report)
        assert len(rows) >= 3  # At least one per scenario

    def test_report_rows_empty(self):
        """_report_rows with empty report should return empty list."""
        report = SimulationReport(results=[], resilience_score=100.0)
        rows = _report_rows(report)
        assert rows == []


class TestReportToExportDict:
    def test_export_dict_structure(self, sample_report: SimulationReport):
        """_report_to_export_dict should produce a valid dict."""
        data = _report_to_export_dict(sample_report)
        assert data["resilience_score"] == 65.0
        assert data["total_scenarios"] == 3
        assert len(data["results"]) == 3

    def test_export_dict_effect_structure(self, sample_report: SimulationReport):
        """Effect dicts should include all expected fields."""
        data = _report_to_export_dict(sample_report)
        # First result has effects
        effects = data["results"][0]["cascade"]["effects"]
        assert len(effects) >= 1
        effect = effects[0]
        assert "component_id" in effect
        assert "component_name" in effect
        assert "health" in effect
        assert "reason" in effect
        assert "estimated_time_seconds" in effect
        assert "metrics_impact" in effect


# ---------------------------------------------------------------------------
# SARIF edge cases
# ---------------------------------------------------------------------------


class TestSarifExportExtended:
    def test_sarif_properties_in_results(self, sample_report: SimulationReport):
        """SARIF results should include properties with risk details."""
        data = json.loads(export_sarif(sample_report))
        results = data["runs"][0]["results"]
        for result in results:
            assert "properties" in result
            assert "risk_score" in result["properties"]
            assert "cascade_severity" in result["properties"]
            assert "affected_components" in result["properties"]

    def test_sarif_message_includes_effects(self, sample_report: SimulationReport):
        """SARIF messages should include affected component details."""
        data = json.loads(export_sarif(sample_report))
        results = data["runs"][0]["results"]
        critical_result = [r for r in results if r["level"] == "error"][0]
        assert "Affected:" in critical_result["message"]["text"]

    def test_sarif_deduplicates_rules(self):
        """Multiple results with same scenario ID should share one rule."""
        scenario = Scenario(
            id="dup-id", name="Dup", description="Dup", faults=[],
        )
        cascade1 = CascadeChain(trigger="Dup", total_components=1)
        cascade1.effects.append(
            CascadeEffect(
                component_id="a", component_name="A",
                health=HealthStatus.DOWN, reason="down",
                estimated_time_seconds=10,
            )
        )
        cascade1.likelihood = 0.8
        cascade2 = CascadeChain(trigger="Dup", total_components=1)
        cascade2.effects.append(
            CascadeEffect(
                component_id="b", component_name="B",
                health=HealthStatus.DOWN, reason="down",
                estimated_time_seconds=10,
            )
        )
        cascade2.likelihood = 0.8

        report = SimulationReport(
            results=[
                ScenarioResult(scenario=scenario, cascade=cascade1, risk_score=8.0),
                ScenarioResult(scenario=scenario, cascade=cascade2, risk_score=8.0),
            ],
            resilience_score=50.0,
        )
        data = json.loads(export_sarif(report))
        rules = data["runs"][0]["tool"]["driver"]["rules"]
        # Should have only 1 rule despite 2 results
        assert len(rules) == 1
        assert len(data["runs"][0]["results"]) == 2
